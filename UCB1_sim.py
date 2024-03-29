#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Chelsea Zackey
UCB1 Bias Simulations
"""
from sympy.stats import Normal as norm, sample
import random
import math
import openpyxl
#import matplotlib.pyplot as plt
#import numpy as np

class Arm:
    def __init__(self, dist):
        self.dist = dist
        self.pulls = 0
        self.rewards = 0
        self.emp_mean = 0
    
    def pull(self, strict=False): #pull arm 
        samp = sample(self.dist)
        if strict==False:
            self.pulls += 1
            self.rewards += samp
            self.emp_mean = float(float(self.rewards) / float(self.pulls))
        return samp
        
    
    def compute_index(self, time): #calculate UCB1 upper bound of arm 
        return self.emp_mean + math.sqrt(2 * math.log(time + 1) / float(self.pulls))
    
    def bias(self, mu): # calculate bias of arm
        return self.emp_mean-mu
        
 
def decide(arms, time): #return index of chosen arm  
    play_arm = 0
    max_index = arms[0].compute_index(time)
    for i in range(1, len(arms)): # calculate and compare all empirical means
        this_index = arms[i].compute_index(time)
        if max_index < this_index:
            max_index = this_index
            play_arm = i
        elif max_index == this_index: #break tie
            tie_break = random.randint(1, 101)
            if tie_break > 50:
                max_index = this_index
                play_arm = i
    return play_arm  

def exploited_arm(arms): #determine which arm bandit exploited most
    max_arm = arms[0].pulls
    exploited = 0
    for i in range(1, len(arms)):
        if max_arm < arms[i].pulls:
            max_arm = arms[i].pulls
            exploited = i
    return exploited   

def bias_stats(arms): #returns dict holding bias of empirical mean of each arm in arms
    biases = []
    for i in range(len(arms)):
        biases.append(arms[i].bias(0.5+i+1))
    return biases


def init_arms(num_arms, arms):
    # creates arms 1-k with each arm i following reward distrbn N(i+0,5, 1)
    for i in range(num_arms-1): # initialize subopt arms with same variance
        arms.append(Arm(norm("Arm"+str(i+1), 0.5+i+1, 1)))
    
    arms.append(Arm(norm("Arm5", 0.5+num_arms, 1))) #change variance of optimal arm here
    

def main():
    sim_runs = 1000 #number of simulation runs
    num_arms = 5 #total number of arms
    data_wkbk = openpyxl.Workbook() # excel workbook to contain data over all simulations 
    r_sheet = data_wkbk.active #regret data sheet
    r_sheet.title = "RegretData"
    b_sheet = data_wkbk.create_sheet("BiasData") #bias data sheet
    p_sheet = data_wkbk.create_sheet("ArmPlaysData") #data sheet for no. times each arm was selected
    
    #format heading for bias data sheet
    for i in range(num_arms):
        b_sheet.cell(row = 1, column = i+1).value = "Arm"+str(i+1)
    #format heading for plays data sheet
    for i in range(num_arms):
        p_sheet.cell(row = 1, column = i+1).value = "Arm"+str(i+1)
    #run each individual simulation
    for i in range(1, sim_runs+1):
        time = num_arms #time counter
        horizon = 500 #simulation horizon
        arms = [] #list of all arms
        reward = 0  # cumulative reward
        regret = 0 # cumulative regret
        bestArm_r = 0
        
        #populate arms[]
        init_arms(num_arms, arms)    
        
        #set title of column of this simulation regret data
        r_sheet.cell(row = 1, column = i).value = "Sim"+str(i)
        
        for k in range(num_arms): #play each arm once
            gain = arms[k].pull()
            reward += gain
            if k == num_arms-1:
                bestArm_r += gain
            else:
                bestArm_r += arms[num_arms-1].pull(strict=True)
                regret = bestArm_r -reward
            r_sheet.cell(row = k+2, column = i).value = float(regret)
        #run rest of simulation
        while(time < horizon): # employ UCB1 algorithm each following round
            choice = decide(arms, time)
            #track cumulative reward & regret
            gain = arms[choice].pull()
            reward += gain
            if choice == num_arms-1:
                bestArm_r += gain
            else:
                bestArm_r += arms[num_arms-1].pull(strict=True)
                regret = bestArm_r -reward
            r_sheet.cell(row = time+2, column = i).value = float(regret)
            time += 1
        
        #output bias stats to excel sheet
        biases = bias_stats(arms) #calculate arm bias stats
        for j in range(num_arms): #print stats
#            b_sheet.cell(row = i+1, column = j+1).value = float(arms[j].emp_mean)
            b_sheet.cell(row = i+1, column = j+1).value = float(biases[j])
       
        #output plays data to excel sheet
        for j in range(num_arms): #print stats
            p_sheet.cell(row = i+1, column = j+1).value = float(arms[j].pulls)
        
            
    #save wkbk data
    data_wkbk.save("/Users/edenzackey/Documents/BanditSimData/simulations.xlsx")
  
        
    #FOR INITIAL TESTING PURPOSES
    
#    #general statistics
#    exploited = exploited_arm(arms)+1 # bandit's selected optimal arm
#    print("===============================Simulation Stats================================")
#    print("")
#    print("Total reward: "+str(reward))
#    print("")
#    print("Total regret: "+str(bestArm_r)
#    print("")
#    print("Most exploited arm: "+str(exploited))
#    print("")
#    
#    # print arm bias stats
#    print("===============================Arm Stats================================")
#    print("")
#    print("Arms and associated biases (on empirical means): ")
#    biases = bias_stats(arms) # bias of empirical means of each arm
#    for j in range(len(biases)):
#        print("Arm "+str(j+1)+": "+str(biases[j]))
#    
#    print("")
#    print("")
#    if exploited == num_arms:
#        print("Bandit successfully determined optimal arm")
#    else: 
#        print("Bandit did not successfully determine optimal arm")
main()    
